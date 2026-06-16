from ollama import ChatResponse
from openpyxl import load_workbook
from ollama import chat
from src.utils import Utils
import pandas as pd
import subprocess
import logging
import json
import time
import os


utils = Utils()


class Agent:
    def __init__(self, data): #data=pd.DataFrame()
        self.model = 'ollama'
        self.data = data


    def call_ollama(self, question):
        role = 'user'
        extract_date = utils.extract_date(question)
        if extract_date:
            response = self.call_simple_response(role, question)
            self._save_agent_response(response, question)
            return response.message.content
        elif utils.is_weekend(str(extract_date)):
                response = f"Please enter a date that doesn't fall on a weekend"
                self._save_fallback_metadata(response, question)
                return response

    def call_simple_response(self, role, question):
        response = chat(
            model='gemma3',
            stream=False,
            messages=[
                {
                    'role': role,
                    'content': f"Here is the data : \n{self.data}/n {question}",
                },
            ],
        )
        return response

    def _save_agent_response(self, response, question):
        meta_data = dict()
        meta_data['timestamp'] = time.time()
        meta_data['logprobs'] = response.logprobs
        meta_data['model'] = str(response.model)
        meta_data['created_at'] = str(response.created_at)
        meta_data['done'] = str(response.done)
        meta_data['done_reason'] = str(response.done_reason)
        meta_data['user_query'] = str(question)
        meta_data['content'] = str(response.message.content)
        meta_data['total_duration'] = str(response.total_duration)
        meta_data['load_duration'] = str(response.load_duration)
        meta_data['prompt_eval_count'] = str(response.prompt_eval_count)
        meta_data['prompt_eval_duration'] = str(response.prompt_eval_duration)
        meta_data['eval_count'] = str(response.eval_count)
        meta_data['eval_duration'] = str(response.eval_duration)
        self.log_metrics(meta_data)

    def log_metrics(self, meta_data):
        file = f"data/metadata.jsonl"
        if os.path.exists(file):
            with open(file, "r") as f:
                try:
                    data = json.load(f)
                except json.JSONDecodeError:
                    data = []
        else:
            data = []
        if not isinstance(data, list):
            data = [data]
        data.append(meta_data)
        with open(file, "w") as f:
            json.dump(data, f, indent=2)

    def _save_fallback_metadata(self, response, question):
        meta_data = dict()
        meta_data['timestamp'] = time.time()
        meta_data['response'] = response
        meta_data['question'] = question
        self.log_metrics(meta_data)
        self.log_metrics(meta_data)


    def _log_prompts(self, user_query, agent_response):
        file = 'prompts/prompts.xlsx'

        if not os.path.exists(file):
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.append(['User Prompt', 'Agent Response'])
            wb.save(file)

        wb = load_workbook(file)
        ws = wb.active

        ws.append([user_query, agent_response])
        wb.save(file)


    # def install_model(self):
    #     """DIY Later"""
    #     capture_output = True
    #     subprocess.call(['curl -fsSL https://ollama.com/install.sh | sh'])
    #     logging.log(msg="Initiating model installation", level=2)
    #     subprocess.call(args='ollama pull gemma3')
    #     subprocess.call(args='ollama run gemma3')



    # def check_ollama_installation(self):
    #     if self.model :
    #         pass
    #     else:
    #         logging.log(msg="Ollama module not installed", level=2)
    #         logging.log(msg="Initiating model installation", level=2)







